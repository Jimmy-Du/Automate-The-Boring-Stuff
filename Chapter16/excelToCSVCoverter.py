# File:        excelToCSVConverter.py
# Description: goes through the current working directory and converts all spreadsheets into csv files. Each sheet in a
#              spreadsheet is given a csv file.



import csv
import openpyxl
import os



# Function:    excelToCSV()
# Description: goes through the specified directory and converts each spreadsheet into a csv file. 1 csv file is created
#              for each sheet in the spreadsheet.
# Parameters:  directory: a path to the directory that contains the excel spreadsheets to be converted to csv files
# Return:      N/A
def excelToCSV(directory):
    # loop to go through all files in the directory specified, checking if they are excel files
    for filename in os.listdir(directory):
        # if the file is an excel spreadsheet, it will be loaded and converted into csv files
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(filename)
            # loop to go through each sheet within the spreadsheet and convert it to a csv file
            for sheetName in wb.get_sheet_names():
                sheet = wb.get_sheet_by_name(sheetName)

                csvFile = open(filename[:-5] + "_" + sheetName + ".csv", "w", newline="")
                csvWriter = csv.writer(csvFile)
                
                # Loop through every row in the sheet.
                for rowNum in range(1, sheet.max_row + 1):
                    rowData = []
                    # Loop through each cell of the row while adding the values in each cell to the row data
                    for colNum in range(1, sheet.max_column + 1):
                        rowData.append(sheet.cell(row=rowNum, column=colNum).value)

                    csvWriter.writerow(rowData)

                csvFile.close()



excelToCSV(".")
