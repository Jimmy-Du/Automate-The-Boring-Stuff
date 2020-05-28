# File:        spreadsheetToTextFile.py
# Description: prompts user to enter the path to the spreadsheet file that is desired to be converted into text files. 
#              Will then load the spreadsheet and loop through the columns and rows adding the contents into 1 or more text
#              files, dependent on how many columns are present. 
#              Each text file created represents 1 column in the spreadsheet, and the file contents of each file is the all the
#              rows within the column.



import openpyxl
import os



# Function:    spreadsheetToTextFile()
# Description: loads the spreadsheet specified and goes through each column and creates a text file containing the
#              contents found in the rows of the column
# Parameters:  file: a string containing the path to the spreadsheet file that will be converted to text files
# Return:      N/A
def spreadsheetToTextFile(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # loop to go through each column and create a file for each column
    for col in range(1, sheet.max_column + 1):
        fp = open(os.path.basename(file) + f"-text{col}.txt", "w")
        # loop to go through each row in the column and write its contents into the newly created file
        for row in range(1, sheet.max_row + 1):
            fp.write(sheet.cell(row=row, column=col).value)

        fp.close()




filePrompt = input("Please enter the path to the spreadsheet file that will be converted to text files:\n")
spreadsheetToTextFile(filePrompt)
