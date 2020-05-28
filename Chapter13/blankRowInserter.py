# File:        blankRowInserter.py
# Description: opens a spreadsheet file specified in the commandline and inserts blank rows into the spreadsheets depenedent
#              on the integer values also passed in on the commandline. 



import sys
import openpyxl



# Function:    insertBlankRows()
# Description: reads in a spreadsheet file and inserts blank rows at the specified locations
# Parameters:  startingRow: an integer indicating the row where the start of the blank rows will be inserted at
#              amountOfRows: an integer indicating the amount of blank rows that will be inserted
#              filename: a string containing the filename of the spreadsheet that will be inserted with blank rows
# Return:      N/A
def insertBlankRows(startingRow, amountOfRows, filename):
    wb  = openpyxl.load_workbook(filename)
    sheet = wb.active

    # nested loops to shift the rows after the inserted blank rows
    for rowToShift in range(sheet.max_row, startingRow, -1):
        for columnToShift in range(1, sheet.max_column + 1):
            sheet.cell(row=rowToShift+1, column=columnToShift).value = sheet.cell(row=rowToShift, column=columnToShift).value
            
    # nested loops to blank out the desired rows
    for blankRow in range(startingRow, startingRow + amountOfRows):
        for blankColumn in range(1, sheet.max_column + 1):
            sheet.cell(row=blankRow, column=blankColumn).value = ''

    wb.save(filename[:-5] + "-blankUpdated.xlsx")



if len(sys.argv) == 4:
    insertBlankRows(int(sys.argv[1]), int(sys.argv[2]), sys.argv[3])
