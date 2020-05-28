# File:        cellInverter.py
# Description: prompts user for the file path to the spreadsheet file to be inverted and goes through the spreadsheet 
#              and changes the rows to columns and columns to rows. 



import openpyxl



# Function:    invertCells()
# Description: reads in a spreadsheet and inverts the cells
# Parameters:  filename: a string containing the filename of the spreadsheet that will have its cells inverted
# Return:      N/A
def invertCells(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # saves the original sheet name to later be used to rename the duplicate sheet created
    sheetTitle = sheet.title
    # creates a new sheet in the spreadsheet to contain the inverted sheet
    duplicateSheet = wb.create_sheet(title="DuplicateSheetForCellInversion")

    # nested loops to inverts the rows and column in the spreadsheet
    for row in range(1, sheet.max_row):
        for col in range(1, sheet.max_column):
            duplicateSheet.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

    # deletes the sheet that was to be updated, then renames the duplicate sheet with inverted cells as the original sheet
    del wb[sheetTitle]
    duplicateSheet.title = sheetTitle

    wb.save(filename[:-5] + "-cellInverted.xlsx")



filePrompt = input("Please enter a the filename for the spreadsheet to be inverted.\n")
invertCells(filePrompt)
