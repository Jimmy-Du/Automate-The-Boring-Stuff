# File:        multiplicationTable.py
# Description: creates a multiplication table in an excel spreadsheet. The max number within the table is determined by 
#              the integer value passed in on the commandline.



import sys
import openpyxl
from openpyxl.styles import Font



# Function:    createMultiTable()
# Description: creates a multiplication table within an excel spreadsheet with the max number of the table determined by the
#              parameter value passed in
# Parameters:  maxNumber: an integer indicating the maximum number that the multiplication table will go up to
# Return:      N/A
def createMultiTable(maxNumber):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # changes the name of the sheet for the table
    sheet.title = f'{maxNumber} x {maxNumber} Table'

    # loop to create the column and row headers for the multiplication table
    for headerCounter in range(1, maxNumber + 1):
        sheet.cell(row=1, column=headerCounter+1).value = headerCounter
        sheet.cell(row=headerCounter+1, column=1).value = headerCounter
        sheet.cell(row=1, column=headerCounter+1).font = Font(bold=True)
        sheet.cell(row=headerCounter+1, column=1).font = Font(bold=True)

    # loop to go through each row of the table filling out the multiplication values
    for row in range(1, maxNumber + 1):
        # loop to go through each column in the row filling out the multiplication values
        for col in range(1, maxNumber + 1):
            sheet.cell(row=row+1, column=col+1).value = row * col

    wb.save('multiplicationTable.xlsx')



# checks that an argument for the max number in the table was provided, if so, the createMultiTable() function is called
if len(sys.argv) == 2:
    try:
        n = int(sys.argv[1])
        createMultiTable(n)
    except:
        print("Argument entered is not a valid integer value.")
